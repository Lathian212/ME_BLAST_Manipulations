""" OUTDATED SEPARATED INTO TWO DIFFERENT MODULES """
""" I wrote this script for Python 3.X """
""" Author: Lathian J.D. Horton (Jonathan S. Kwiat) Date: 7-09-16
    This uses the Biopython's XML parser. The package is installed by
    $pip install biopython. The entire setup is in a virtual environment
    provided by python 3.x called pyvenv. This allows the biopython
    package to be installed into it and 'frozen' so any changes to biopython
    will not break the code here.

    The purpose of this program is to take the blastn query-hit records
    that are stored in one large xml file
    (OUTPUT_CII-MWP0002.contigs_singletons.blastx.viral_Queries.xml)
    created by doing a local blastn (with just defaults) query against a copy
    of the TSA Leptopilina heterotoma databse and breaking this file down
    into individual records, files.

    Also the best hit for each query will be placed into the already existing
    Blastn_viral_queries_against_Lh_TSA.xlsx file. This file was created
    by the make_qorksheet_queries.py script. It was needed because the return
    from blastn does NOT INCLUDE THE ORIGINAL QUERY JUST THE HSPS.
"""




""" Bio.SearchIO is still experimental at this point will throw warning flag,
    which only means the authors consider it changeable but the virtual
    evironment will protect against any changes assuming nobody upgrades
    the bioPython in virtualPython3.5
"""
from Bio import SearchIO
""" Decimal class used to get consistent precision from evalue float for file 
    naming"""
from decimal import *
""" To create, erase directories, and get list of directory contents os, shutil,
    subprocess modules need to be imported"""
import os
import subprocess
import shutil
# To get information on floats minimum
import sys
""" To work with Excel .xlsx files I am using openpyxl """
from openpyxl import Workbook
from openpyxl import load_workbook
# all functions I define will be here
def outputFileName (query_count, hit, hsp):
    """This function will return a string for the filename of an output file.
       Q_[QueryIDName]_H_[HitAccesionName]_[Evalue].xml evalue set to 2
       decimal places of percision"""
    getcontext().prec = 2
    return str('Query_' + str(query_count) + '_H_' + hit.accession + '_' + \
               str (Decimal (hsp.evalue) / Decimal (1)) + '.xml')
def wipe_dir(fDic):
    """ For the purposes of debugging this script I define this function
        to first wipe the directories created by a previous run of the program 
        and then create the directories anew because logical errors from the
        previous run can result in inappropiate files staying in folders they 
        do not belong. This function will just create appropiate folders 
        if there is no previous directory 
    """
    if os.path.exists(fDic['topDir']):
        ans = input('The Directory Tree Structure for output already exists' +
                    ' enter Y/N to erase directory.')
        if ans.upper() == 'Y':
            shutil.rmtree(fDic['topDir'])
    if not os.path.exists(fDic['topDir']) :
        # https://docs.python.org/2/library/os.html#files-and-directories
        os.makedirs(fDic['topDir'])
        os.makedirs(fDic['topDir'] + fDic['noHit'])
def ask_wipe_excel_file():
    """ Asks for an Excel file name to save summary information and then
        Tests whether the Excel summary file already exists and if so deletes
        the file.
    """
    fileName = input('Please enter a file name without the xlsx file ' + \
                     'extension to create a summary Excel file.')
    if os.path.isfile(fileName + '.xlsx'):
        ans = input('An Excel worksheet of this name is already in the ' +
                    'current directory. Erase it and start over?(Y/N)')
    if ans.upper() == 'Y':
            os.remove(fileName + '.xlsx')
            print('Deleting old summary file of the same name and '+
                  'recreating it.')
    return fileName
def set_cell(ws, r, c, val):
    """ Sets a cell given the work_sheet, row, column, and value. """
    current_cell = ws.cell(row=r , column=c)
    current_cell.value = val
    c += 1
    return c
    
def build_ws_header(work_sheet, max_hits):
    """ Writes out to the worksheet header identifiers. """
    first_header_info = ['Query #', 'Query Sequence',
                        'Top Hit Accession in L.h.', 'E-Value', 'Filename']
    r = 1
    c = 1
    for val in first_header_info:
        c = set_cell(ws, r, c, val)
def write_hit_to_file(query_count, query_reslt, hit_num, work_sheet):
    pass
def write_query_sequence(ws_output, ws_input):
    pass
def write_qr_to_ws(query_count, query_result, work_sheet):
    """ Writes out the qr, query_result, to the Excel summary sheet in the
        following manner. Query # Total Hits, Hit #, Total HSPS, Best Hsp,
        Lowest Evalue for hit. Calls SearchIO.write function and makes
        hyperlinks to those files in the Hit # column or the Toal Hits column if
        there are zero hits.
    """
    qr = query_result
    ws = work_sheet
    c = 1
    # query_count holds the info of what row the data needs to be written to
    r = 1 + query_count
    # set_cell advances column by one
    # set Query_#
    c = set_cell(ws, r , c, ('Query_' + str(query_count)))
    # set Total Hits
    tot_hits = len(qr)
    c = set_cell(ws, r , c, tot_hits)
    
    # NOTE this is duplicated code from split function so could be refactored
    # 'hits' inside SearchIO class is a list the starts at zero
    hit_num = 0
    for hit in query_result.hits:
        # Set hit # col; Hit 1 Hit 2 etc
        c = set_cell(ws, r, c, (hit_num + 1))
        # Get and Set accession id of the hit.
        c = set_cell(ws, r, c, (hit.accession))
        # Get total number hsps for hit and set it to worksheet.
        total_hsps = len (hit.hsps)
        c = set_cell(ws, r, c, (total_hsps))
        # counter to keep track of which hsp is 'best' for the hit
        hsp_num = 0
        lowest_eval = hit.hsps[0].evalue
        for hsp in hit.hsps:
            if hsp.evalue < lowest_eval:
                lowest_eval = hsp.evalue
                best_hsp_num = hsp_num
        # Set 'Best High Scoring Pair' starint numbering at 1
            hsp_num += 1
        c = set_cell(ws, r, c, (hsp_num))
        # Set lowest evalue for best HSP before going on to next hit
        c = set_cell(ws, r, c, lowest_eval)
    
def ask_directory_name():
    """ This ask the user to select the top directory name and calls wipe_dir
        to ask user if they wish to wipe away a previous output run 
        of the program
    """
    topDirVal = input('Please enter the name you wish the top directory ' + 
                    'to have (XXXX_Split): ') + '_Split/'
    print ("Thank you.")
    # https://docs.python.org/2/tutorial/datastructures.html#dictionaries
    # Setting up a python dictionary, a key : value pair so all the directory path information
    # is stored in one place and can be constructed by fDic['KEY1'] + fDic['Key2']
    fDic = {'topDir' : topDirVal , 'noHit' : 'no_Hits_at_Any_E/'}
    # Below function checks if top directory is already there and offers to wipe it if so.
    wipe_dir(fDic)
    return fDic
def start_queryResult_generator(inFile, fDic, work_sheet):
    """ invoking the parse function to return a 'generator' that can allow you 
        to step though the record one QueryResult Object at a time but invoking
        nextQuery = (next)generator on it.This approach can allow you to save 
        on memory. I have found with my current task casting this generator with
        (list) works fine but it is really not called for in this current 
        task of parsing and sorting the records.
    """
    """ http://biopython.org/DIST/docs/api/Bio.SearchIO.BlastIO-module.html"""
    qGenerator = SearchIO.parse(inFile, 'blast-xml')
    max_hits = 0
    query_count = 1
    # Step through all the records in the lump xml data file and write out
    # each separate hit to file. Also write the summary information to the
    # work sheet.
    for query_result in qGenerator:
        print('Processing Query BLAST return ' + str(query_count))
        number_hits = int(len(query_result.hits))
        # Extend header out right if new MAXHITS
        if number_hits > max_hits:
            max_hits = number_hits       
        if number_hits == 0:
            # Construct path plus file name for no hit query
            filename = str(fDic['topDir'] + fDic['noHit'] + 'Query_' 
                           + str(query_count) + '_H_none.xml')
            # Write out any Queries that had to hits to a no Hit subfolder
            SearchIO.write(query_result, filename, 'blast-xml')
            write_qr_to_ws(query_count, query_result, work_sheet)
        else :
            # Now set up a counter of 'hits' in the QueryResult so hit's
            # can be sliced away into their own record cleanly.
            hit_count = 0;
            for hit in query_result.hits:
                total_hsps = len (hit.hsps)
                lowest_eval = hit.hsps[0].evalue
                best_hsp = hit.hsps[0]
                for hsp in hit.hsps:
                    if hsp.evalue < lowest_eval:
                        lowest_eval = hsp.evalue
                        best_hsp = hsp
                filename = str(fDic['topDir'] + outputFileName(query_count, hit, best_hsp))
                SearchIO.write(query_result[hit_count:(hit_count + 1)], filename , 'blast-xml')
                hit_count += 1
            # Write out query_result to worksheet           
            write_qr_to_ws(query_count, query_result, work_sheet)
        query_count += 1
        # break is debugging code
        # if query_count == 20:
        #   break
    build_ws_header(work_sheet, max_hits)
    return qGenerator
# End of Functions start of 'main' program
""" This program will now search the directory it was started in for files
    ending in XML and then ask the User if they would like to
    Split the file into individual records and create an
    Excel workseet that holds summary and relative hyperlinks to each file.

    Please note the following is a call to system command 'ls' that only exists
    in MacOSx and Linux. If you run this script in Windows you need 
    the 'dir' command to replace it.
"""
"""
    suprocess.run with the PIPE connected to stdout and with universal_newlines
    set to True returns a subprocess.CompletedProcess object that has the
    stdout stored in subprocess.CompletedProcess.stdout
"""
# For debugging the evalue see the min float value then should I set it to zero?
# print(sys.float_info)
excelFileName = ask_wipe_excel_file()
# Create a new empty Excel WorkBook in memory.
wb = Workbook()
# Get the automatically created worksheet
ws = wb.active
# Set title of worksheet to excelFileName
ws.title = excelFileName
# Look for file ending in xml to split
output = subprocess.run(['ls'], stdout=subprocess.PIPE, universal_newlines=True)
output = output.stdout.split()
for val in output:
    if '.xml' in val:
        ans = input('The program found this file: ' + val + \
               ' do you want to Split it into individual records? (Y/N): ')
        if ans.upper() == 'Y':
            fDic = ask_directory_name()
            qGenerator = start_queryResult_generator(val, fDic, ws)
wb.save(excelFileName + '.xlsx')
"""
print("Program Finished")
"""
