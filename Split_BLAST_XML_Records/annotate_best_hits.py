'''
Created on Jul 10, 2016

@author: jonathan
The purpose of this script is to take the file:
best_hits_Blastn_viral_queries_against_Lh_TSA.xlsx
and merge in the annotations from the blastx results from the Lipkin lab in:
CII-MWP0002.contigs_singletons.blastx.viral_report.xlsx
'''
""" To work with Excel .xlsx files I am using openpyxl """
from openpyxl import Workbook
from openpyxl import load_workbook
#MAIN STARTS
#Most of the report has 30 columns and it does in the header
wb_report = load_workbook(
    'CII-MWP0002.contigs_singletons.blastx.viral_report.xlsx', read_only= False)
ws_report = wb_report.active
# This workbook will be saved under another name with merge in filename
wb_best_hit = load_workbook(
    'best_hits_Blastn_viral_queries_against_Lh_TSA.xlsx', read_only= False)
# To perserve starting input best hit workbook save it to another file
wb_best_hit.save('merged.xlsx')
wb_merge = load_workbook('merged.xlsx')
ws_merge = wb_merge.active
merge_row = 2
merge_col = 6
# Put in header from report (just as an exercise)
for i in range(1,31):
    val = ws_report.cell(row = 1 , column = i).value
    merge_cell = ws_merge.cell(row = 2, column = (6 + i -1))
    merge_cell.value = val
# Range of query sequences in best_hits file
for i in range (3, 4292):
    best_hit_query = ws_merge.cell(row = i, column = 2).value
    # Range of query seequence rows in viral report.
    for j in range (2, 4291):
        report_val = ws_report.cell(row = j, column = 20).value
        if best_hit_query == report_val:
            #Note columns 22 to 30 are empty after row 3993 so this if is needed
            if j <= 3993 :
                for k in range (1, 31):
                    report_val = ws_report.cell(row = j, column = k).value
                    query_cell = ws_merge.cell(row = i, column = k +5)
                    query_cell.value = report_val
            else:
                for k in range (1, 22):
                    report_val = ws_report.cell(row = j, column = k).value
                    query_cell = ws_merge.cell(row = i, column = k +5)
                    query_cell.value = report_val
wb_merge.save('Blastn_against_Lh_TSA_merged_Lipkin_Viral_Report.xlsx')
print('Program done.')