'''
Created on Jul 9, 2016

@author: jonathan

I realized that the task of adding the best hit for each blastn viral query
against the Leptopilina hetertoma TSA to the:
Blastn_viral_queries_againstLh_TSA.xlsx Spreadsheet 
and the task of slicing the lump xml results in:
OUTPUT_CII-MWP0002.contigs_singletons.blsatx.viral_Queries.xml
needed to be separated.

This is the module for adding the best hit to the Excel spreadsheet.
'''
""" To work with Excel .xlsx files I am using openpyxl """
from openpyxl import Workbook
from openpyxl import load_workbook
""" Bio.SearchIO is still experimental at this point will throw warning flag,
    which only means the authors consider it changeable but the virtual
    evironment will protect against any changes assuming nobody upgrades
    the bioPython in virtualPython3.5
"""
from Bio import SearchIO

#MAIN STARTS
wb_queries = load_workbook('C_Blastn_viral_queries_against_Lh_TSA.xlsx', 
                           read_only= False)
ws = wb_queries.active
"""Note starting row and col will change if more data added to ws"""
ws_row = 3
"""will vary between 3 and 5 in range"""
ws_col = 3
blastn = 'Results_blastn_of_blastx_viral_Queries_against_Lh_TSA.xml'
qGenerator = SearchIO.parse(blastn, 'blast-xml')
for query_result in qGenerator:
        #print('Processing Query BLAST return ' + str(ws_row-2))
        number_hits = int(len(query_result.hits))
        # write number of hits to 3 column and appropriate row.
        curr_cell = ws.cell(row = ws_row, column = ws_col)
        curr_cell.value = number_hits
        if number_hits == 0:
            ws_row += 1
            ws_col = 3
            continue
        ws_col += 1
        curr_cell = ws.cell(row = ws_row, column = ws_col)
        best_acc = query_result.hits[0].accession
        curr_cell.value = best_acc
        ws_col += 1
        curr_cell = ws.cell(row = ws_row, column = ws_col)
        best_eval = query_result.hits[0].hsps[0].evalue
        curr_cell.value = best_eval
        ws_row += 1
        ws_col = 3
        if ws_row == 4292:
            break
outfile = 'best_hits_Blastn_viral_queries_against_Lh_TSA.xlsx'
wb_queries.save(outfile)
print('program done')

