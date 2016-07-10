'''
Created on Jul 9, 2016
This is a little script to take the xlsx query file Mary Ellen gave me and
turn out Query_# Query Sequence
@author: jonathan
'''
from openpyxl import Workbook
from openpyxl import load_workbook

wb1 = load_workbook('CII-MWP0002.contigs_singletons.blastx.viral_Queries.xlsx',
                   read_only=True, use_iterators= True)
ws1 = wb1.active
wb2 = Workbook()
ws2 = wb2.active
ws2.title = 'V Queries BLASTn to L.h. TSA'
h2_1 = ws2.cell(row =1 , column =1)
h2_1.value = 'Query #'
h2_2 = ws2.cell(row =1 , column =2)
h2_2.value = 'Query Sequence' 
#Now transfer Query Sequences to new worksheet
query_num = 1
for row in ws1.iter_rows(row_offset = 0):
    for cell in row:
        #Remove whitespace
        temp = cell.value.replace(" ", "")
        print(str(query_num) + '   ' +  temp)
        ws2_col1 = ws2.cell(row = (query_num + 1), column = 1)
        ws2_col2 = ws2.cell(row = (query_num + 1), column = 2)
        ws2_col1.value = 'Query_' + str(query_num)
        ws2_col2.value = temp
        query_num += 1
wb2.save('Blastn_viral_queries_against_Lh_TSA.xlsx')
print('Program done.')